# Importa o Pandas e NumPy.
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import re

# RELATÓRIO
def gerar_n_relatorio():
    # Lê as planilhas.
    nartic = pd.read_excel('Relatório.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'FILIAL_R', 'FILIAL', 'BLOQ_NARTIC', 'QTD_VENDAS']
    nartic = nartic.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    nartic['PEDIDO_QUANT'] = pd.NA
    nartic['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    nartic = nartic[re_ordem]

    # Cria a planilha formatada.
    nartic.to_excel('Nartic Relatório.xlsx', index=False)

# PEDIDO
def gerar_n_pedido():
    # Lê as planilhas.
    nartic = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'FILIAL_R', 'FILIAL', 'BLOQ_NARTIC', 'QTD_VENDAS']
    nartic = nartic.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    nartic['PEDIDO_QUANT'] = pd.NA
    nartic['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    nartic = nartic[re_ordem]

    # Exclui vazios.
    nartic = nartic.dropna(subset=['CODIGO', 'PALLET', 'NARTIC'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    nartic = nartic.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    nartic['NARTIC'] = nartic['NARTIC'].astype(int)
    nartic['PALLET'] = nartic['PALLET'].astype(int)
    nartic['PEDIDO_QUANT'] = nartic['PEDIDO_QUANT'].astype(int)
    nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    nartic['LOCALIZACAO'] = nartic['LOCALIZACAO'].astype(str)
    nartic['DESCRICAO'] = nartic['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    resultado_moeda = nartic[nartic[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    resultado_moeda.to_excel('Nartic Pedido.xlsx', index=False) #Cria uma nova planilha, 'index=False' exlcui o ID criado pelo pandas.

# ZERO ESTOQUE
def gerar_n_zero_estoque():
    # Lê as planilhas.
    nartic = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'FILIAL_R', 'FILIAL', 'BLOQ_NARTIC', 'QTD_VENDAS']
    nartic = nartic.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    nartic['PEDIDO_QUANT'] = pd.NA
    nartic['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    nartic = nartic[re_ordem]

    # Exclui vazios.
    nartic = nartic.dropna(subset=['CODIGO', 'PALLET', 'NARTIC'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    nartic = nartic.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    nartic['NARTIC'] = nartic['NARTIC'].astype(int)
    nartic['PALLET'] = nartic['PALLET'].astype(int)
    nartic['PEDIDO_QUANT'] = nartic['PEDIDO_QUANT'].astype(int)
    nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    nartic['LOCALIZACAO'] = nartic['LOCALIZACAO'].astype(str)
    nartic['DESCRICAO'] = nartic['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'

    resultado_nartic = nartic[nartic[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    resultado_transf = transf[transf[colunas_comum].isin(nartic[colunas_comum])]
    transf = resultado_transf

    # Tratamento na coluna de "CODIGO" na transferência para valores duplicados.
    transf = transf.groupby('CODIGO', as_index=False)['QUANTIDADE'].max()

    # Lê a planilha.
    nartic = resultado_nartic

    nartic = nartic.sort_values(by='CODIGO').reset_index(drop=True)
    transf = transf.sort_values(by='CODIGO').reset_index(drop=True)

    # Explicação:
    # sort_values(by='CODIGO'): Ordena as planilhas pela coluna CODIGO (ou qualquer outra coluna que você escolher).
    # reset_index(drop=True): Reseta os índices das planilhas para garantir que o índice seja sequencial e contínuo após a ordenação. O parâmetro drop=True impede que o índice antigo seja mantido como uma nova coluna.

    nartic['PEDIDO_QUANT'] = transf['QUANTIDADE']
    nartic['PEDIDO_QUANT'] = nartic['NARTIC'].where(nartic['NARTIC'] < transf['QUANTIDADE'], other=transf['QUANTIDADE'])
    nartic['PEDIDO_QUANT'] = transf['QUANTIDADE'].where(nartic['PALLET_QUANT'] == 0, other=transf['QUANTIDADE'])
    nartic['PALLET_QUANT'] = np.floor((nartic['PEDIDO_QUANT'] / nartic['PALLET'])* 10) / 10

    # nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].round().astype(float) #round() pega a metade e joga para cima ou para baixo 'ceil()' joga tudo para cima.

    # Pega os zerados e separa.
    nartic_zero_estoque = nartic[nartic['NARTIC'] == 0]

    # Exclui os zerados
    nartic = nartic[nartic['NARTIC'] != 0]

    nartic = nartic.sort_values(by='DESCRICAO').reset_index(drop=True)

    nartic_zero_estoque.to_excel('Nartic Zero Estoque.xlsx', index=False)

# CÓDIGO ERRADO
def gerar_n_cod_errado():
    # Lê as planilhas.
    nartic = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'FILIAL_R', 'FILIAL', 'BLOQ_NARTIC', 'QTD_VENDAS']
    nartic = nartic.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    nartic['PEDIDO_QUANT'] = pd.NA
    nartic['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    nartic = nartic[re_ordem]

    # Exclui vazios.
    nartic = nartic.dropna(subset=['CODIGO', 'PALLET', 'NARTIC'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    nartic = nartic.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    nartic['NARTIC'] = nartic['NARTIC'].astype(int)
    nartic['PALLET'] = nartic['PALLET'].astype(int)
    nartic['PEDIDO_QUANT'] = nartic['PEDIDO_QUANT'].astype(int)
    nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    nartic['LOCALIZACAO'] = nartic['LOCALIZACAO'].astype(str)
    nartic['DESCRICAO'] = nartic['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    resultado_transf_zero = transf[~transf[colunas_comum].isin(nartic[colunas_comum])]
    resultado_transf_zero.to_excel('Relatório Código Errado Nartic.xlsx', index=False)

# FINALIZADO
def gerar_n_finalizado():
    # Lê as planilhas.
    nartic = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'FILIAL_R', 'FILIAL', 'BLOQ_NARTIC', 'QTD_VENDAS']
    nartic = nartic.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    nartic['PEDIDO_QUANT'] = pd.NA
    nartic['PALLET_QUANT'] = pd.NA
    nartic['QUANT_CX'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']
    nartic = nartic[re_ordem]

    # Exclui vazios.
    nartic = nartic.dropna(subset=['CODIGO', 'PALLET', 'NARTIC'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    nartic = nartic.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    nartic['NARTIC'] = nartic['NARTIC'].astype(int)
    nartic['PALLET'] = nartic['PALLET'].astype(int)
    nartic['PEDIDO_QUANT'] = nartic['PEDIDO_QUANT'].astype(int)
    nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    nartic['LOCALIZACAO'] = nartic['LOCALIZACAO'].astype(str)
    nartic['DESCRICAO'] = nartic['DESCRICAO'].astype(str)
    nartic['QUANT_CX'] = nartic['QUANT_CX'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    resultado_nartic = nartic[nartic[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    resultado_transf = transf[transf[colunas_comum].isin(nartic[colunas_comum])]
    transf = resultado_transf

    # Tratamento na coluna de "CODIGO" na transferência para valores duplicados.
    transf = transf.groupby('CODIGO', as_index=False)['QUANTIDADE'].max()

    # Lê a planilha.
    nartic = resultado_nartic

    nartic = nartic.sort_values(by='CODIGO').reset_index(drop=True)
    transf = transf.sort_values(by='CODIGO').reset_index(drop=True)

    # Explicação:
    # sort_values(by='CODIGO'): Ordena as planilhas pela coluna CODIGO (ou qualquer outra coluna que você escolher).
    # reset_index(drop=True): Reseta os índices das planilhas para garantir que o índice seja sequencial e contínuo após a ordenação. O parâmetro drop=True impede que o índice antigo seja mantido como uma nova coluna.

    nartic['PEDIDO_QUANT'] = transf['QUANTIDADE']
    nartic['PEDIDO_QUANT'] = nartic['NARTIC'].where(nartic['NARTIC'] < transf['QUANTIDADE'], other=transf['QUANTIDADE'])
    nartic['PEDIDO_QUANT'] = transf['QUANTIDADE'].where(nartic['PALLET_QUANT'] == 0, other=transf['QUANTIDADE'])
    nartic['PALLET_QUANT'] = np.floor((nartic['PEDIDO_QUANT'] / nartic['PALLET'])* 10) / 10

    # nartic['PALLET_QUANT'] = nartic['PALLET_QUANT'].round().astype(float) #round() pega a metade e joga para cima ou para baixo 'ceil()' joga tudo para cima.

    # Exclui os zerados
    nartic = nartic[nartic['NARTIC'] != 0]

    nartic = nartic.sort_values(by='DESCRICAO').reset_index(drop=True)

    # Cria uma coluna com número da linha excel.
    nartic['linha_excel'] = nartic.index +2

    # Aplica a fórmula na coluna QUANT_CX.
    nartic['QUANT_CX'] = nartic.apply(lambda row: gerar_formula_excel(row['DESCRICAO'], row['linha_excel']), axis=1)

    # Remove a coluna auxiliar.
    nartic.drop(columns=['linha_excel'], inplace=True)

    nartic.to_excel('Nartic Finalizado.xlsx', index=False)

    # Ler a planilha.
    wb = load_workbook('Nartic Finalizado.xlsx')
    ws = wb.active

    for col in ws.iter_cols(1, ws.max_column):
        nm_col = col[0].value

        # Fonte e tamanho.
        if nm_col in ['CODIGO', 'DESCRICAO', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']:
            for linha in col:
                linha.font = Font(name='Arial', size=10)
        elif nm_col in ['REFFOR']:
            for linha in col:
                linha.font = Font(name='Arial', size=8)

        # Centralização.
        if nm_col in ['CODIGO', 'REFFOR', 'CODVOL', 'NARTIC', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']:
            for linha in col:
                linha.alignment = Alignment(horizontal='center', vertical='center')

        # Negrito e cor.
        if nm_col in ['CODIGO', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']:
            for linha in col:
                linha.font = Font(bold=True)
        elif nm_col in ['NARTIC', 'PALLET']:
            for linha in col:
                linha.font = Font(bold=True, color='FF0000')

    # Salvar.
    wb.save('Nartic Finalizado.xlsx')

def gerar_formula_excel(texto, linha):
    # Confere se tem "CX" para não criar fórmula inútil.
    if re.search(r'cx', str(texto), re.IGNORECASE):
        return f'=(G{linha}/VALUE(MID(C{linha}, FIND("CX", C{linha}) +2, 10))) & "cx"'
    elif re.search(r'fd', str(texto), re.IGNORECASE):
        return f'=(G{linha}/VALUE(MID(C{linha}, FIND("FD", C{linha}) +2, 10))) & "fd"'
    else:
        return ''