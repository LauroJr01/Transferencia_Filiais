# Importa o Pandas e NumPy.
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import re

# RELATÓRIO
def gerar_m_relatorio():
    # Lê as planilhas.
    moeda = pd.read_excel('Relatório.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'NARTIC', 'BLOQ_NARTIC', 'QTD_VENDAS']
    moeda = moeda.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    moeda['PEDIDO_QUANT'] = pd.NA
    moeda['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    moeda = moeda[re_ordem]

    # Cria a planilha formatada.
    moeda.to_excel('Moeda Relatório.xlsx', index=False)

# PEDIDO
def gerar_m_pedido():
    # Lê as planilhas.
    moeda = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'NARTIC', 'BLOQ_NARTIC', 'QTD_VENDAS']
    moeda = moeda.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    moeda['PEDIDO_QUANT'] = pd.NA
    moeda['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    moeda = moeda[re_ordem]

    # Exclui vazios.
    moeda = moeda.dropna(subset=['CODIGO', 'PALLET', 'FILIAL'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    moeda = moeda.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    moeda['FILIAL'] = moeda['FILIAL'].astype(int)
    moeda['PALLET'] = moeda['PALLET'].astype(int)
    moeda['PEDIDO_QUANT'] = moeda['PEDIDO_QUANT'].astype(int)
    moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    moeda['LOCALIZACAO'] = moeda['LOCALIZACAO'].astype(str)
    moeda['DESCRICAO'] = moeda['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    m_pedido = moeda[moeda[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    m_pedido.to_excel('Moeda Pedido.xlsx', index=False) #Cria uma nova planilha, 'index=False' exlcui o ID criado pelo pandas.

# ZERO ESTOQUE
def gerar_m_zero_estoque():
    # Lê as planilhas.
    moeda = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'NARTIC', 'BLOQ_NARTIC', 'QTD_VENDAS']
    moeda = moeda.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    moeda['PEDIDO_QUANT'] = pd.NA
    moeda['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    moeda = moeda[re_ordem]

    # Exclui vazios.
    moeda = moeda.dropna(subset=['CODIGO', 'PALLET', 'FILIAL'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    moeda = moeda.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    moeda['FILIAL'] = moeda['FILIAL'].astype(int)
    moeda['PALLET'] = moeda['PALLET'].astype(int)
    moeda['PEDIDO_QUANT'] = moeda['PEDIDO_QUANT'].astype(int)
    moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    moeda['LOCALIZACAO'] = moeda['LOCALIZACAO'].astype(str)
    moeda['DESCRICAO'] = moeda['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'

    resultado_moeda = moeda[moeda[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais
    resultado_transf = transf[transf[colunas_comum].isin(moeda[colunas_comum])]
    transf = resultado_transf

    # Tratamento na coluna de "CODIGO" na transferência para valores duplicados.
    transf = transf.groupby('CODIGO', as_index=False)['QUANTIDADE'].max()

    # Lê a planilha.
    moeda = resultado_moeda

    moeda = moeda.sort_values(by='CODIGO').reset_index(drop=True)
    transf = transf.sort_values(by='CODIGO').reset_index(drop=True)

    # Explicação:
    # sort_values(by='CODIGO'): Ordena as planilhas pela coluna CODIGO (ou qualquer outra coluna que você escolher).
    # reset_index(drop=True): Reseta os índices das planilhas para garantir que o índice seja sequencial e contínuo após a ordenação. O parâmetro drop=True impede que o índice antigo seja mantido como uma nova coluna.

    moeda['PEDIDO_QUANT'] = transf['QUANTIDADE']
    moeda['PEDIDO_QUANT'] = moeda['FILIAL'].where(moeda['FILIAL'] < transf['QUANTIDADE'], other=transf['QUANTIDADE'])
    moeda['PEDIDO_QUANT'] = transf['QUANTIDADE'].where(moeda['PALLET_QUANT'] == 0, other=transf['QUANTIDADE'])
    moeda['PALLET_QUANT'] = np.floor((moeda['PEDIDO_QUANT'] / moeda['PALLET'])* 10) / 10

    #moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].round().astype(float) #round() pega a metade e joga para cima ou para baixo 'ceil()' joga tudo para cima.

    # Pega os zerados e separa.
    moeda_zero_estoque = moeda[moeda['FILIAL'] == 0]

    # Exclui os zerados
    moeda = moeda[moeda['FILIAL'] != 0]

    moeda = moeda.sort_values(by='DESCRICAO').reset_index(drop=True)

    moeda_zero_estoque.to_excel('Moeda Zero Estoque.xlsx', index=False)

# CÓDIGO ERRADO
def gerar_m_cod_errado():
    # Lê as planilhas.
    moeda = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'NARTIC', 'BLOQ_NARTIC', 'QTD_VENDAS']
    moeda = moeda.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    moeda['PEDIDO_QUANT'] = pd.NA
    moeda['PALLET_QUANT'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']
    moeda = moeda[re_ordem]

    # Exclui vazios.
    moeda = moeda.dropna(subset=['CODIGO', 'PALLET', 'FILIAL'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    moeda = moeda.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    moeda['FILIAL'] = moeda['FILIAL'].astype(int)
    moeda['PALLET'] = moeda['PALLET'].astype(int)
    moeda['PEDIDO_QUANT'] = moeda['PEDIDO_QUANT'].astype(int)
    moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    moeda['LOCALIZACAO'] = moeda['LOCALIZACAO'].astype(str)
    moeda['DESCRICAO'] = moeda['DESCRICAO'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    relatorio_cod_errado = transf[~transf[colunas_comum].isin(moeda[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    relatorio_cod_errado.to_excel('Relatório Código Errado Moeda.xlsx', index=False) #Cria uma nova planilha, 'index=False' exlcui o ID criado pelo pandas.

# FINALIZADO
def gerar_m_finalizado():
    # Lê as planilhas.
    moeda = pd.read_excel('Relatório.xlsx')
    transf = pd.read_excel('Transferência.xlsx')

    # Exclui colunas desnecessárias.
    colunas_drop = ['MATRIZ', 'NARTIC', 'BLOQ_NARTIC', 'QTD_VENDAS']
    moeda = moeda.drop(columns=colunas_drop)

    # Reorganiza as colunas.
    moeda['PEDIDO_QUANT'] = pd.NA
    moeda['PALLET_QUANT'] = pd.NA
    moeda['QUANT_CX'] = pd.NA
    re_ordem = ['CODIGO', 'REFFOR', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']
    moeda = moeda[re_ordem]

    # Exclui vazios.
    moeda = moeda.dropna(subset=['CODIGO', 'PALLET', 'FILIAL'])
    transf = transf.dropna(subset=['CODIGO', 'QUANTIDADE'])

    # Preenche valores vazios.
    moeda = moeda.fillna(0)
    transf = transf.fillna(0)

    # Formata a planilha (arruma os tipos).
    moeda['FILIAL'] = moeda['FILIAL'].astype(int)
    moeda['PALLET'] = moeda['PALLET'].astype(int)
    moeda['PEDIDO_QUANT'] = moeda['PEDIDO_QUANT'].astype(int)
    moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].astype(int)
    transf['CODIGO'] = transf['CODIGO'].astype(int)
    transf['QUANTIDADE'] = transf['QUANTIDADE'].astype(int)
    moeda['LOCALIZACAO'] = moeda['LOCALIZACAO'].astype(str)
    moeda['DESCRICAO'] = moeda['DESCRICAO'].astype(str)
    moeda['QUANT_CX'] = moeda['QUANT_CX'].astype(str)

    # Pega colunas em comum.
    colunas_comum = 'CODIGO'
    resultado_moeda = moeda[moeda[colunas_comum].isin(transf[colunas_comum])] #'isin' Faz o comparativo entre ambas colunas, excluindo as demais.
    resultado_transf = transf[transf[colunas_comum].isin(moeda[colunas_comum])]
    transf = resultado_transf

    # Tratamento na coluna de "CODIGO" na transferência para valores duplicados.
    transf = transf.groupby('CODIGO', as_index=False)['QUANTIDADE'].max()

    # Lê a planilha.
    moeda = resultado_moeda

    moeda = moeda.sort_values(by='CODIGO').reset_index(drop=True)
    transf = transf.sort_values(by='CODIGO').reset_index(drop=True)

    # Explicação:
    # sort_values(by='CODIGO'): Ordena as planilhas pela coluna CODIGO (ou qualquer outra coluna que você escolher).
    # reset_index(drop=True): Reseta os índices das planilhas para garantir que o índice seja sequencial e contínuo após a ordenação. O parâmetro drop=True impede que o índice antigo seja mantido como uma nova coluna.

    moeda['PEDIDO_QUANT'] = transf['QUANTIDADE']
    moeda['PEDIDO_QUANT'] = moeda['FILIAL'].where(moeda['FILIAL'] < transf['QUANTIDADE'], other=transf['QUANTIDADE'])
    moeda['PEDIDO_QUANT'] = transf['QUANTIDADE'].where(moeda['PALLET_QUANT'] == 0, other=transf['QUANTIDADE'])
    moeda['PALLET_QUANT'] = np.floor((moeda['PEDIDO_QUANT'] / moeda['PALLET'])* 10) / 10

    #moeda['PALLET_QUANT'] = moeda['PALLET_QUANT'].round().astype(float) #round() pega a metade e joga para cima ou para baixo 'ceil()' joga tudo para cima.

    # Exclui os zerados.
    moeda = moeda[moeda['FILIAL'] != 0]

    moeda = moeda.sort_values(by='DESCRICAO').reset_index(drop=True)

    # Cria uma coluna com número da linha excel.
    moeda['linha_excel'] = moeda.index +2

    # Aplica a fórmula na coluna QUANT_CX.
    moeda['QUANT_CX'] = moeda.apply(lambda row: gerar_formula_excel(row['DESCRICAO'], row['linha_excel']), axis=1)
                    
    # Remove a coluna auxiliar.
    moeda.drop(columns=['linha_excel'], inplace=True)

    moeda.to_excel('Moeda Finalizado.xlsx', index=False)

    # Ler a planilha.
    wb = load_workbook('Moeda Finalizado.xlsx')
    ws = wb.active

    for col in ws.iter_cols(1, ws.max_column):
        nm_col = col[0].value

        # Fonte e tamanho.
        if nm_col in ['CODIGO', 'DESCRICAO', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']:
            for linha in col:
                linha.font = Font(name='Arial', size=10)
        elif nm_col in ['REFFOR']:
            for linha in col:
                linha.font = Font(name='Arial', size=8)

        # Centralização.
        if nm_col in ['CODIGO', 'REFFOR', 'CODVOL', 'FILIAL_R', 'FILIAL', 'PALLET', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO', 'QUANT_CX']:
            for linha in col:
                linha.alignment = Alignment(horizontal='center', vertical='center')

        # Negrito e cor.
        if nm_col in ['CODIGO', 'PEDIDO_QUANT', 'PALLET_QUANT', 'LOCALIZACAO']:
            for linha in col:
                linha.font = Font(bold=True)
        elif nm_col in ['FILIAL', 'PALLET']:
            for linha in col:
                linha.font = Font(bold=True, color='FF0000')

    # Salvar.
    wb.save('Moeda Finalizado.xlsx')

def gerar_formula_excel(texto, linha):
    # Confere se tem "CX" para não criar fórmula inútil.
    if re.search(r'cx', str(texto), re.IGNORECASE):
        return f'=(H{linha}/VALUE(MID(C{linha}, FIND("CX", C{linha}) +2, 10))) & "cx"'
    elif re.search(r'fd', str(texto), re.IGNORECASE):
        return f'=(H{linha}/VALUE(MID(C{linha}, FIND("FD", C{linha}) +2, 10))) & "fd"'
    else:
        return ''